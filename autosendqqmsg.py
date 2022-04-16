from ast import arg
# from asyncio import threads
import imp
import re
import win32con
import win32gui
import win32clipboard as w
import time
from random import choice
from random import randrange
import datetime
from openpyxl import load_workbook
import _thread
import threading
import sys
import os

class sendMsg():
    def __init__(self,receiver,msg):
        self.receiver=receiver
        self.msg=msg
    
    #设置剪贴版内容
    def setText(self):
        w.OpenClipboard()
        w.EmptyClipboard()
        w.SetClipboardData(win32con.CF_UNICODETEXT, self.msg)
        w.CloseClipboard()
    #发送消息
    def sendmsg(self):
        time1=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        qq=win32gui.FindWindow(None,self.receiver)
        win32gui.SendMessage(qq,win32con.WM_PASTE , 0, 0)
        win32gui.SendMessage(qq, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
        print(time1+" sucessfuly send",self.msg)


class myexcel():
    def __init__(self,fileName):
        self.fileName=fileName
        self.excel=load_workbook(filename=fileName)["Sheet1"]
    #获取对象名字
    def getreceiver(self):
        return self.excel['A2'].value
    #获取最大行列数
    def get_row_num(self):
        rows = self.excel.max_row
        # columns = self.ws.max_column
        return rows
    #获取任务
    def gettasks(self,row):
        data = []
        for i in self.excel[row]:
            data.append(i.value)
        return data


def getmessage(fileName):
    f=open(fileName,'r',encoding='utf-8')
    lines=f.readlines()
    f.close()
    return choice(lines)


def task(data,receiver):

    if(int(data[2])!=0):
        print('佳人们，小睡一会')
        time.sleep(int(data[2]))
    while True:
        msg=data[0]
        qq=sendMsg(receiver,msg)
        qq.setText()
        time.sleep(3)
        qq.sendmsg()
        time.sleep(int(data[1]))
    

def main():

    # print (os.path.abspath('..'))
    excel=myexcel('./Desktop/qq/dom.XLSX')
    receiver=excel.getreceiver()
    rows=excel.get_row_num()
    mythreads=[]
    for i in range(rows-2):
        data=excel.gettasks(i+3)
        a=threading.Thread(target=task,args=(data,receiver))
        a.setDaemon(True)
        mythreads.append(a)
        a.start()
        print("线程"+str(i+1)+"启动")
  
    while True:
        alive=False
        for i in range(rows-2):
            alive = alive or mythreads[i].is_alive()
        if not alive:
            break
    
    # while True:
    #     time1=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #     # print(time1)
    #     hour=time1[11:13]
    #     minnute = time1[14:16]
    #     second = time1[17:19]
    #     # print(hour,minnute,second)
    
       
    #     # if int(hour)%2==0 and minnute=='00' and second=='30':
    #     if second=='30':
    #         counter = counter+1
    #         print(time1)
    #         # msg = getmessage('C:\\Users\\Administrator\\Desktop\\message.txt')
    #         msg1 = '探索五行草原'
    #         msg2 = '探索上古战场'
    #         msg3 = '探索生灵禁区'
    #         qq1 = sendMsg(receiver, msg1)
    #         qq1.sendmsg()
    #         if counter == 3:
    #             counter = 0;
    #             qq2 = sendMsg(receiver,msg2)
    #             qq2.sendmsg()
    #         # print(time1+':'+msg2)
        
    #             qq3 = sendMsg(receiver,msg3)
    #             qq3.sendmsg()
    #         # print(time1+':'+msg1)
        
    #         time.sleep(10)


        # 测试案例  每5秒钟发送一次消息
        # if int(second)%5==0:
        #     msg = '时间:' + time1 + ',系统运行正常！'
        #     qq = sendMsg(receiver, msg)
        #     qq.sendmsg()
        #     time.sleep(3)

if __name__ == '__main__':
    main()
